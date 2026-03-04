"""
Lógica de conciliação SIGE × Mercado Pago.
Recebe os bytes das duas planilhas e retorna os bytes do Excel de saída.
"""
import io

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ── utilidades ────────────────────────────────────────────────────────────────

def _norm_cols(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = [str(c).strip().upper() for c in df.columns]
    return df


def _find_col(df: pd.DataFrame, *candidates) -> str | None:
    """Retorna a primeira coluna que casar (exato, depois parcial)."""
    cols = df.columns.tolist()
    for cand in candidates:
        cu = cand.upper()
        if cu in cols:
            return cu
    for cand in candidates:
        cu = cand.upper()
        for c in cols:
            if cu in c or c in cu:
                return c
    return None


def _find_mp_header(raw: bytes) -> int:
    """Detecta a linha do cabeçalho real na exportação do MP."""
    df = pd.read_excel(io.BytesIO(raw), header=None, nrows=30)
    for i, row in df.iterrows():
        vals = {str(v).strip().upper() for v in row if pd.notna(v)}
        if {"CREDITADO", "DEBITADO"} <= vals:
            return int(i)
    return 0


def _safe(v):
    """Retorna None para NaN/None, caso contrário o float."""
    if v is None:
        return None
    try:
        f = float(v)
        return None if np.isnan(f) else f
    except (TypeError, ValueError):
        return None


def _clean_id(v) -> str | None:
    s = str(v).strip() if pd.notna(v) else ""
    return s if s not in ("", "nan", "None") else None


# ── função principal ──────────────────────────────────────────────────────────

def processar(sige_bytes: bytes, mp_bytes: bytes) -> bytes:
    # ── 1. Lê SIGE ────────────────────────────────────────────────────────
    df_s = _norm_cols(pd.read_excel(io.BytesIO(sige_bytes)))

    ci = _find_col(df_s, "ID VENDA MERCADO LIVRE", "ID VENDA ML", "ID ML", "ID")
    cc = _find_col(df_s, "CLIENTE", "NOME CLIENTE", "NOME")
    cv = _find_col(df_s, "VALOR", "VALOR TOTAL", "TOTAL")

    if not all([ci, cc, cv]):
        raise ValueError(
            f"Planilha SIGE: colunas não encontradas.\n"
            f"Esperado: ID / CLIENTE / VALOR\n"
            f"Encontradas: {df_s.columns.tolist()}"
        )

    df_s = df_s[pd.notna(df_s[ci])].copy()
    df_s[ci] = df_s[ci].astype(str).str.strip()
    df_s[cv] = pd.to_numeric(df_s[cv], errors="coerce")

    sige: dict[str, dict] = {}
    for _, row in df_s.iterrows():
        sid = row[ci]
        if sid and sid not in ("nan", "None", ""):
            sige[sid] = {"cliente": row[cc], "valor": _safe(row[cv])}

    # ── 2. Lê Mercado Pago ────────────────────────────────────────────────
    h = _find_mp_header(mp_bytes)
    df_m = _norm_cols(pd.read_excel(io.BytesIO(mp_bytes), header=h))

    c_data = _find_col(df_m, "DATA", "DATA DA OPERAÇÃO", "DATA OPERACAO", "DATA CRIAÇÃO")
    c_desc = _find_col(df_m, "DESCRIÇÃO", "DESCRICAO", "DESCRIPTION")
    c_ref  = _find_col(df_m, "CÓDIGO DE REFERÊNCIA", "CODIGO DE REFERENCIA", "REFERÊNCIA", "REFERENCIA")
    c_ped  = _find_col(df_m, "ID DO PEDIDO", "ID PEDIDO", "PEDIDO")
    c_pac  = _find_col(df_m, "ID DO PACOTE", "ID PACOTE", "PACOTE")
    c_op   = _find_col(df_m, "ID DA OPERAÇÃO NO MERCADO PAGO", "ID DA OPERACAO", "ID OPERAÇÃO", "ID OPERACAO")
    c_cred = _find_col(df_m, "CREDITADO")
    c_deb  = _find_col(df_m, "DEBITADO")
    c_sal  = _find_col(df_m, "SALDO")

    if not c_cred or not c_deb:
        raise ValueError(
            f"Planilha MP: CREDITADO/DEBITADO não encontrados.\n"
            f"Encontradas: {df_m.columns.tolist()}"
        )

    # Captura saldo ANTES de remover as linhas de rodapé
    sal_vals = df_m[c_sal].dropna().tolist() if c_sal else []
    saldo_ini = _safe(sal_vals[0])  if len(sal_vals) >= 1 else None
    # penúltima = última linha antes de "Total"
    saldo_fin = _safe(sal_vals[-2]) if len(sal_vals) >= 2 else _safe(sal_vals[-1]) if sal_vals else None

    # Remove linhas de cabeçalho/rodapé do MP
    if c_desc:
        bad = df_m[c_desc].astype(str).str.strip().str.lower().isin(
            ["saldo inicial disponível", "total"]
        )
        df_m = df_m[~bad].copy()

    # Normaliza numéricos
    df_m[c_cred] = pd.to_numeric(df_m[c_cred], errors="coerce").fillna(0)
    df_m[c_deb]  = pd.to_numeric(df_m[c_deb],  errors="coerce").fillna(0)
    if c_data:
        df_m[c_data] = pd.to_datetime(df_m[c_data], errors="coerce")
    for col in filter(None, [c_ref, c_ped, c_pac, c_op]):
        df_m[col] = df_m[col].apply(_clean_id)

    # ── 3. Chave de agrupamento ───────────────────────────────────────────
    def group_key(row) -> str:
        for col in filter(None, [c_pac, c_ped, c_ref]):
            v = row.get(col)
            if v:
                return v
        op = row.get(c_op) if c_op else None
        return f"__op__{op}" if op else f"__row__{row.name}"

    df_m["_gk"] = df_m.apply(group_key, axis=1)

    # ── 4. Agrega grupos ──────────────────────────────────────────────────
    rows_matched:   list[dict] = []
    rows_unmatched: list[dict] = []

    for gk, grp in df_m.groupby("_gk", sort=False):
        cred = grp[c_cred].sum()
        deb  = grp[c_deb].sum()
        cld  = cred - deb

        if abs(cld) < 0.001:       # saldo zero → descarta
            continue

        data = grp[c_data].min() if c_data else None
        if data is not None and pd.isna(data):
            data = None

        # Tenta casar com SIGE
        info = None
        for col in filter(None, [c_ref, c_ped, c_pac]):
            for v in grp[col].dropna().unique():
                if v and v in sige:
                    info = sige[v]
                    break
            if info:
                break

        if info:
            # ID CERTO: ref → pedido → pacote
            id_certo = None
            for col in filter(None, [c_ref, c_ped, c_pac]):
                vals = [v for v in grp[col].dropna().unique() if v]
                if vals:
                    id_certo = vals[0]
                    break
            id_certo = id_certo or str(gk)

            vs     = info["valor"]
            tarifa = (vs - cld) if vs is not None else None
            rows_matched.append({
                "DATA":       data,
                "ID CERTO":   id_certo,
                "CLIENTE":    info["cliente"],
                "CREDITADO":  cred,
                "DEBITADO":   deb,
                "CMD":        cld,
                "VALOR SIGE": vs,
                "TARIFA":     tarifa,
            })
        else:
            # Sem match → agrupa por ID DA OPERAÇÃO
            descs = []
            if c_desc:
                descs = list(dict.fromkeys(
                    str(v).strip() for v in grp[c_desc].dropna()
                    if str(v).strip() not in ("", "nan", "None")
                ))
            gk_clean = str(gk).replace("__op__", "").replace("__row__", "")
            rows_unmatched.append({
                "DATA":       data,
                "ID CERTO":   gk_clean,
                "CLIENTE":    ", ".join(descs) if descs else None,
                "CREDITADO":  cred,
                "DEBITADO":   deb,
                "CMD":        cld,
                "VALOR SIGE": cld,
                "TARIFA":     0.0,
            })

    # Ordena matched por VALOR SIGE crescente; unmatched sempre ao final
    rows_matched.sort(key=lambda r: r["VALOR SIGE"] if r["VALOR SIGE"] is not None else 0.0)

    total_vs     = sum(r["VALOR SIGE"] for r in rows_matched if r["VALOR SIGE"] is not None)
    total_tarifa = sum(r["TARIFA"]     for r in rows_matched if r["TARIFA"]     is not None)

    # ── 5. Monta o Excel ──────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Reconciliação"

    FILL_HDR  = PatternFill("solid", fgColor="1F4E79")
    FILL_YEL  = PatternFill("solid", fgColor="FFF2CC")
    FILL_BLU  = PatternFill("solid", fgColor="D6E4F0")
    FILL_GRN  = PatternFill("solid", fgColor="E2EFDA")
    FONT_H    = Font(name="Arial", size=10, bold=True,  color="FFFFFF")
    FONT_N    = Font(name="Arial", size=10, bold=False, color="000000")
    FONT_B    = Font(name="Arial", size=10, bold=True,  color="000000")
    NUM       = "#,##0.00;[Red]-#,##0.00"
    DATE_FMT  = "DD/MM/YYYY"

    HEADERS = ["DATA", "ID CERTO", "CLIENTE", "CREDITADO", "DEBITADO",
               "CREDITADO MENOS DEBITADO", "VALOR SIGE", "TARIFA"]
    WIDTHS  = [12, 20, 45, 14, 14, 22, 14, 14]

    # Cabeçalho
    ws.row_dimensions[1].height = 30
    for ci, (h_txt, w) in enumerate(zip(HEADERS, WIDTHS), 1):
        cell = ws.cell(row=1, column=ci, value=h_txt)
        cell.fill      = FILL_HDR
        cell.font      = FONT_H
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = w

    def write_row(ri: int, r: dict, fill=None):
        vals = [
            r["DATA"], r["ID CERTO"], r["CLIENTE"],
            r["CREDITADO"], r["DEBITADO"], r["CMD"],
            r["VALOR SIGE"], r["TARIFA"],
        ]
        for ci, v in enumerate(vals, 1):
            if isinstance(v, float) and np.isnan(v):
                v = None
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.font = FONT_N
            if fill:
                cell.fill = fill
            if ci == 1:
                cell.number_format = DATE_FMT
            elif ci >= 4:
                cell.number_format = NUM

    ri = 2
    for r in rows_matched:
        write_row(ri, r)
        ri += 1
    for r in rows_unmatched:
        write_row(ri, r, fill=FILL_YEL)
        ri += 1

    # Linha TOTAL
    for ci in range(1, 9):
        cell = ws.cell(row=ri, column=ci)
        cell.fill = FILL_BLU
        cell.font = FONT_B
    ws.cell(row=ri, column=3, value="TOTAL")
    cell_vs = ws.cell(row=ri, column=7, value=total_vs)
    cell_vs.number_format = NUM
    cell_tf = ws.cell(row=ri, column=8, value=total_tarifa)
    cell_tf.number_format = NUM
    ri += 2  # linha em branco

    # Saldo inicial / final
    for label, val in [("SALDO INICIAL", saldo_ini), ("SALDO FINAL", saldo_fin)]:
        ws.cell(row=ri, column=1).fill = FILL_GRN
        lbl = ws.cell(row=ri, column=2, value=label)
        lbl.fill      = FILL_GRN
        lbl.font      = FONT_B
        lbl.alignment = Alignment(horizontal="right")
        v_cell = ws.cell(row=ri, column=3, value=val)
        v_cell.fill = FILL_GRN
        v_cell.font = FONT_B
        if val is not None:
            v_cell.number_format = NUM
        ri += 1

    # Oculta colunas D, E, F
    for col in ["D", "E", "F"]:
        ws.column_dimensions[col].hidden = True

    ws.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
